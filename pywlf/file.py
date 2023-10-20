import os
import json
import shutil
import chardet
import openpyxl
import hashlib


def mod_encoding(filepath: str, dst_encoding: str) -> None:
    """修改文件编码

    Args:
        filepath (str): 文件路径
        dst_encoding (str): 目标编码
    """
    src_encoding = get_encoding(filepath)
    file_data = ""
    with open(filepath, "r", encoding=src_encoding) as f:
        for line in f:
            file_data += line
    with open(filepath, "w", encoding=dst_encoding) as f:
        f.write(file_data)


def get_encoding(filepath: str) -> str:
    """获取文件编码

    Args:
        filepath (str): 文件路径

    Returns:
        str: 文件编码
    """
    with open(filepath, "rb") as fp:
        return chardet.detect(fp.read(1024 * 1024))["encoding"]


def mod_str(filepath: str, lns: list, old_str: str, new_str: str) -> None:
    """替换文件中所有指定行的所有指定字符

    Args:
        filepath (str): 文件路径
        lns (list): 行号，指定的所有行都替换
        old_str (str): 旧字符串
        new_str (str): 新字符串
    """
    encoding = get_encoding(filepath)
    file_data, ln = "", 0
    with open(filepath, "r", encoding=encoding) as f:
        for line in f:
            ln += 1
            if ln in lns:
                if old_str in line:
                    line = line.replace(old_str, new_str)
            file_data += line
    with open(filepath, "w", encoding=encoding) as f:
        f.write(file_data)


def list_dir(filepath: str, contains: str = None) -> list:
    """获取目录下的所有文件

    Args:
        filepath (str): 文件路径
        contains (str, optional): 文件路径包含的字符. Defaults to None.

    Returns:
        list: 目录下的所有文件
    """
    res = []
    for file in os.listdir(filepath):
        fp = f'{filepath}/{file}'
        if contains:
            if contains in file:
                res.append(fp)
        else:
            res.append(fp)
    res.sort()
    return res


def new_dir_ifn(filepath: str) -> None:
    """创建目录，不存在则创建，存在无操作

    Args:
        filepath (str): 文件路径
    """
    if not os.path.exists(filepath):
        os.makedirs(filepath)


def del_fd(filepath: str) -> None:
    """删除文件或目录

    Args:
        filepath (str): 文件路径
    """
    if os.path.isdir(filepath):
        shutil.rmtree(filepath)
    elif os.path.isfile(filepath):
        os.remove(filepath)


def file_size(filepath: str) -> str:
    """获取文件或文件夹的大小

    Args:
        filepath (str): 文件路径

    Returns:
        str: 文件或文件夹的大小
    
    Ps:
        超过TB的数据就别用了，需要考虑性能了
    """
    # 判断输入是文件夹还是文件
    if os.path.isdir(filepath):
        # 如果是文件夹则统计文件夹下所有文件的大小
        for file in os.listdir(filepath):
            res += os.path.getsize(f'{filepath}/{file}')
    elif os.path.isfile(filepath):
        # 如果是文件则直接统计文件的大小
        res += os.path.getsize(filepath)
    # 格式化返回大小
    bu = 1024
    if res < bu:
        res = f'{bu}B'
    elif bu <= res < bu**2:
        res = f'{round(res / bu, 3)}KB'
    elif bu**2 <= res < bu**3:
        res = f'{round(res / bu**2, 3)}MB'
    elif bu**3 <= res < bu**4:
        res = f'{round(res / bu**3, 3)}GB'
    elif bu**4 <= res < bu**5:
        res = f'{round(res / bu**4, 3)}TB'
    else:
        res = 'NaN'
    return res


def cls_dir(filepath: str) -> None:
    """清空文件夹

    Args:
        filepath (str): 文件路径
    """
    if os.path.exists(filepath):
        shutil.rmtree(filepath)
    os.mkdir(filepath)


def from_json(filepath: str, k1: str = None, k2: str = None, k3: str = None) -> any:
    """读json文件，返回数组或字典结构数据

    Args:
        filepath (str): 文件路径
        k1 (str, optional): 一级深度key. Defaults to None.
        k2 (str, optional): 二级深度key. Defaults to None.
        k3 (str, optional): 三级深度key. Defaults to None.

    Returns:
        list | dict: 返回数组或字典结构数据
    """
    with open(filepath, "r", encoding='utf-8') as fr:
        data = json.load(fr)
    data = eval(data) if type(data) is str else data
    if k1:
        if k2:
            if k3:
                v1 = get_key(data, k1)
                v2 = get_key(v1, k2)
                res = get_key(v2, k3)
            else:
                v1 = get_key(data, k1)
                res = get_key(v1, k2)
        else:
            res = get_key(data, k1)
    else:
        res = data
    return res


def get_key(data: any, key: str) -> any:
    """从json数据中获取指定的key

    Args:
        data (any): json数据
        key (str): key

    Returns:
        dict|list|NaN: 返回key的值
    """
    if type(data) is dict:
        res = data.get(key)
    elif type(data) is list:
        res = [e.get(key) for e in data]
    else:
        res = None
    return res


def to_json(filepath: str, data: any, cn=True, encoding='utf8') -> None:
    """把数据写入到json文件中

    Args:
        filepath (str): 文件路径
        data (list | dict): 数组或字典结果数据
        cn (bool, optional): 是否支持中文. Defaults to True.
        encoding (str, optional): 文件编码. Defaults to 'utf8'.
    """
    json_data = json.dumps(data, ensure_ascii=not cn)
    with open(filepath, "w", encoding=encoding) as fw:
        fw.write(json_data)


def from_xlsx(filepath: str, sheet_name: str = 'Sheet1') -> list:
    """从excel文件中读数据

    Args:
        filepath (str): 文件路径
        sheet_name (str, optional): sheet名. Defaults to 'Sheet1'.

    Returns:
        list: 返回每行数据的集合
    """
    res = []
    wk = openpyxl.Workbook(filepath)
    sheet = wk[sheet_name]
    for i in range(1, sheet.max_row + 1):
        rows = []
        for j in range(1, sheet.max_column + 1):
            rows.append(sheet.cell(row=i, column=j).value)
        res.append(rows)
    return res


def to_xlsx(filepath: str, data: list, sheet_name: str = None, autoh=False, headers: list = None) -> None:
    """把数据写入到excel文件中

    Args:
        filepath (str): 文件路径
        data (list): 数组或字典结构数据
        autoh (bool, optional): 自动添加表头，字典的key. Defaults to False.
        headers (list, optional): 表头. Defaults to None.
    """
    wk = openpyxl.Workbook()
    sheet = wk[sheet_name] if sheet_name else wk.active
    if type(data[0]) is list:
        if headers:
            data.insert(0, headers)
        for i, lv in enumerate(data):
            for j, v in enumerate(lv):
                sheet.cell(row=i + 1, column=j + 1).value = v
    elif type(data[0]) is dict:
        if headers:
            rc = {k: k for k in headers}
            data.insert(0, rc)
        if autoh and not headers:
            rc = {k: k for k in data[0]}
            data.insert(0, rc)
        for i, rc in enumerate(data):
            for j, k in enumerate(rc):
                sheet.cell(row=i + 1, column=j + 1).value = rc[k]
    wk.save(filepath)


def md5(filepath: str) -> str:
    """获取文件的哈希值MD5

    Args:
        filepath (str): 文件路径

    Returns:
        str: 哈希值MD5
    """
    m = hashlib.md5()
    with open(filepath, "rb") as fr:
        while True:
            data = fr.read(1024)
            if not data:
                break
            m.update(data)
    return m.hexdigest()
